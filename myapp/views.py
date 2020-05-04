import csv  # per export in csv
import xlwt  # per export in xsl
from openpyxl import Workbook  # per export in xslx
from openpyxl.styles import Font, Alignment  # per lo stile delle celle
from openpyxl.utils import get_column_letter  # per lavorare con le colonne
from django.http import HttpResponse
from .models import User
from django.shortcuts import render


def index(request):
    return render(request, "index.html")


# Export dati in formato csv
def export_users_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="users.csv"'

    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Username', 'First name', 'Last name', 'Email address'])

    users = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
    for user in users:
        writer.writerow(user)

    return response


# Export dati in formato xls
def export_users_xls(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="users.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Users')

    # Indice dell'header
    row_num = 0

    style = xlwt.XFStyle()
    style.font.bold = True

    columns = ['Username', 'First name', 'Last name', 'Email address', ]

    # Scrittura dell'header
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], style)

    # Righe rimanenti
    # contenenti i dati estratti dalla tabella
    style = xlwt.XFStyle()

    # Imposto i bordi
    borders = xlwt.Borders()
    borders.left = 0
    borders.right = 0
    borders.top = 2
    borders.bottom = 2

    style.borders = borders

    rows = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], style)

    wb.save(response)
    return response


# Export dati in formato xlsx
# https://openpyxl.readthedocs.io/en/stable/
# https://djangotricks.blogspot.com/2019/02/how-to-export-data-to-xlsx-files.html
def export_users_xlsx(request):
    # Queryset dati
    users = User.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    response['Content-Disposition'] = 'attachment; filename=users.xlsx'

    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'users'

    # Definizione titoli colonne
    columns = ['Username', 'First name', 'Last name', 'Email address', ]

    row_num = 1

    # Assegno il titolo ad ogni colonna
    for col_num, column_title in enumerate(columns, 1):
        # imposto la larghezza della colonna
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 40

        cell = worksheet.cell(row=row_num, column=col_num)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(name='Calibri', bold=True)
        cell.value = column_title

    for user in users:
        row_num += 1

        # Recupero dati per ogni riga
        row = [
            user.username,
            user.first_name,
            user.last_name,
            user.email,
        ]

        # Scrittura riga nel foglio
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response
